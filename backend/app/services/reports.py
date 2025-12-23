"""
Servicio de Reportes y Exportación
"""
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func
from io import BytesIO
from datetime import datetime
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.chart import RadarChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models import (
    Company, Area, User, Question, Category,
    QuestionnaireAssignment, QuestionAssignment, Response,
    DivergenceAlert, AlertSeverity
)


class ReportService:
    """Servicio para generación de reportes"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def get_dashboard_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas generales del dashboard"""
        total_companies = self.db.query(Company).filter(Company.is_active == True).count()
        total_users = self.db.query(User).filter(User.is_active == True).count()
        total_questions = self.db.query(Question).filter(Question.is_active == True).count()
        
        active_questionnaires = self.db.query(QuestionnaireAssignment).filter(
            QuestionnaireAssignment.is_active == True
        ).count()
        
        total_assignments = self.db.query(QuestionAssignment).count()
        completed_responses = self.db.query(Response).count()
        pending_responses = total_assignments - completed_responses
        
        divergence_alerts = self.db.query(DivergenceAlert).filter(
            DivergenceAlert.is_resolved == False
        ).count()
        
        completion_rate = (completed_responses / total_assignments * 100) if total_assignments > 0 else 0
        
        return {
            "total_companies": total_companies,
            "total_users": total_users,
            "total_questions": total_questions,
            "active_questionnaires": active_questionnaires,
            "pending_responses": pending_responses,
            "completed_responses": completed_responses,
            "divergence_alerts": divergence_alerts,
            "completion_rate": round(completion_rate, 2)
        }
    
    def get_company_report(self, company_id: int, questionnaire_id: Optional[int] = None) -> Dict[str, Any]:
        """Generar reporte completo para una empresa"""
        company = self.db.query(Company).filter(Company.id == company_id).first()
        if not company:
            return None
        
        # Query base
        base_query = self.db.query(
            Area.id.label("area_id"),
            Area.name.label("area_name"),
            func.sum(Response.score).label("total_score"),
            func.sum(Question.max_score).label("max_score"),
            func.count(Response.id).label("questions_answered"),
            func.count(QuestionAssignment.id).label("total_questions")
        ).select_from(QuestionAssignment).join(
            Question, QuestionAssignment.question_id == Question.id
        ).join(
            User, QuestionAssignment.user_id == User.id
        ).join(
            Area, User.area_id == Area.id
        ).outerjoin(
            Response, QuestionAssignment.id == Response.assignment_id
        ).join(
            QuestionnaireAssignment, QuestionAssignment.questionnaire_id == QuestionnaireAssignment.id
        ).filter(
            QuestionnaireAssignment.company_id == company_id
        )
        
        if questionnaire_id:
            base_query = base_query.filter(QuestionnaireAssignment.id == questionnaire_id)
        
        areas_data = base_query.group_by(Area.id, Area.name).all()
        
        # Calcular scores por área
        areas_scores = []
        total_score = 0
        total_max = 0
        
        for area in areas_data:
            score = float(area.total_score or 0)
            max_score = float(area.max_score or 0)
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            areas_scores.append({
                "area_id": area.area_id,
                "area_name": area.area_name,
                "score": round(score, 2),
                "max_score": round(max_score, 2),
                "percentage": round(percentage, 2),
                "questions_answered": area.questions_answered,
                "total_questions": area.total_questions
            })
            
            total_score += score
            total_max += max_score
        
        # Scores por categoría
        categories_query = self.db.query(
            Category.id.label("category_id"),
            Category.name.label("category_name"),
            Category.color.label("category_color"),
            func.sum(Response.score).label("total_score"),
            func.sum(Question.max_score).label("max_score")
        ).select_from(QuestionAssignment).join(
            Question, QuestionAssignment.question_id == Question.id
        ).join(
            Category, Question.category_id == Category.id
        ).outerjoin(
            Response, QuestionAssignment.id == Response.assignment_id
        ).join(
            QuestionnaireAssignment, QuestionAssignment.questionnaire_id == QuestionnaireAssignment.id
        ).filter(
            QuestionnaireAssignment.company_id == company_id
        )
        
        if questionnaire_id:
            categories_query = categories_query.filter(QuestionnaireAssignment.id == questionnaire_id)
        
        categories_data = categories_query.group_by(Category.id, Category.name, Category.color).all()
        
        categories_scores = []
        for cat in categories_data:
            score = float(cat.total_score or 0)
            max_score = float(cat.max_score or 0)
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            categories_scores.append({
                "category_id": cat.category_id,
                "category_name": cat.category_name,
                "category_color": cat.category_color,
                "score": round(score, 2),
                "max_score": round(max_score, 2),
                "percentage": round(percentage, 2)
            })
        
        # Alertas de divergencia
        alerts_query = self.db.query(DivergenceAlert).join(
            QuestionnaireAssignment
        ).filter(
            QuestionnaireAssignment.company_id == company_id
        )
        
        if questionnaire_id:
            alerts_query = alerts_query.filter(DivergenceAlert.questionnaire_id == questionnaire_id)
        
        divergence_alerts = []
        for alert in alerts_query.all():
            question = self.db.query(Question).filter(Question.id == alert.question_id).first()
            divergence_alerts.append({
                "id": alert.id,
                "question_text": question.text if question else "",
                "severity": alert.severity.value,
                "responses_data": alert.responses_data,
                "variance": alert.variance,
                "is_resolved": alert.is_resolved,
                "created_at": alert.created_at.isoformat()
            })
        
        overall_percentage = (total_score / total_max * 100) if total_max > 0 else 0
        
        return {
            "company_id": company_id,
            "company_name": company.name,
            "overall_score": round(total_score, 2),
            "overall_percentage": round(overall_percentage, 2),
            "areas_scores": areas_scores,
            "categories_scores": categories_scores,
            "divergence_alerts": divergence_alerts
        }
    
    def export_to_excel(self, company_id: int, questionnaire_id: Optional[int] = None) -> BytesIO:
        """Exportar datos a Excel"""
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Obtener datos
        company = self.db.query(Company).filter(Company.id == company_id).first()
        
        # Hoja 1: Resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        
        ws_summary["A1"] = "REPORTE DE CUMPLIMIENTO"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A2"] = f"Empresa: {company.name}"
        ws_summary["A3"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        report = self.get_company_report(company_id, questionnaire_id)
        
        ws_summary["A5"] = "PUNTAJE GENERAL"
        ws_summary["A5"].font = Font(bold=True)
        ws_summary["A6"] = f"{report['overall_percentage']}%"
        ws_summary["A6"].font = Font(size=24, bold=True)
        
        # Tabla de áreas
        ws_summary["A8"] = "Cumplimiento por Área"
        ws_summary["A8"].font = Font(bold=True)
        
        headers = ["Área", "Puntaje", "Máximo", "Porcentaje", "Respondidas", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=9, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border
        
        for row, area in enumerate(report["areas_scores"], 10):
            ws_summary.cell(row=row, column=1, value=area["area_name"]).border = border
            ws_summary.cell(row=row, column=2, value=area["score"]).border = border
            ws_summary.cell(row=row, column=3, value=area["max_score"]).border = border
            ws_summary.cell(row=row, column=4, value=f"{area['percentage']}%").border = border
            ws_summary.cell(row=row, column=5, value=area["questions_answered"]).border = border
            ws_summary.cell(row=row, column=6, value=area["total_questions"]).border = border
        
        # Ajustar anchos
        for col in range(1, 7):
            ws_summary.column_dimensions[chr(64 + col)].width = 15
        
        # Hoja 2: Respuestas detalladas
        ws_responses = wb.create_sheet("Respuestas")
        
        # Query de todas las respuestas
        responses_query = self.db.query(
            User.full_name.label("user_name"),
            User.email.label("user_email"),
            Area.name.label("area_name"),
            Category.name.label("category_name"),
            Question.code.label("question_code"),
            Question.text.label("question_text"),
            Question.question_type,
            Response.answer,
            Response.score,
            Response.answered_at
        ).select_from(QuestionAssignment).join(
            User, QuestionAssignment.user_id == User.id
        ).join(
            Area, User.area_id == Area.id
        ).join(
            Question, QuestionAssignment.question_id == Question.id
        ).outerjoin(
            Category, Question.category_id == Category.id
        ).join(
            Response, QuestionAssignment.id == Response.assignment_id
        ).join(
            QuestionnaireAssignment, QuestionAssignment.questionnaire_id == QuestionnaireAssignment.id
        ).filter(
            QuestionnaireAssignment.company_id == company_id
        )
        
        if questionnaire_id:
            responses_query = responses_query.filter(QuestionnaireAssignment.id == questionnaire_id)
        
        # Headers
        response_headers = ["Usuario", "Email", "Área", "Categoría", "Código", "Pregunta", "Tipo", "Respuesta", "Puntaje", "Fecha"]
        for col, header in enumerate(response_headers, 1):
            cell = ws_responses.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border
        
        for row, resp in enumerate(responses_query.all(), 2):
            answer_str = json.dumps(resp.answer) if isinstance(resp.answer, (dict, list)) else str(resp.answer)
            
            ws_responses.cell(row=row, column=1, value=resp.user_name).border = border
            ws_responses.cell(row=row, column=2, value=resp.user_email).border = border
            ws_responses.cell(row=row, column=3, value=resp.area_name).border = border
            ws_responses.cell(row=row, column=4, value=resp.category_name or "").border = border
            ws_responses.cell(row=row, column=5, value=resp.question_code or "").border = border
            ws_responses.cell(row=row, column=6, value=resp.question_text).border = border
            ws_responses.cell(row=row, column=7, value=resp.question_type.value).border = border
            ws_responses.cell(row=row, column=8, value=answer_str).border = border
            ws_responses.cell(row=row, column=9, value=resp.score or 0).border = border
            ws_responses.cell(row=row, column=10, value=resp.answered_at.strftime('%d/%m/%Y %H:%M') if resp.answered_at else "").border = border
        
        # Ajustar anchos
        ws_responses.column_dimensions['A'].width = 20
        ws_responses.column_dimensions['B'].width = 25
        ws_responses.column_dimensions['C'].width = 15
        ws_responses.column_dimensions['D'].width = 15
        ws_responses.column_dimensions['E'].width = 10
        ws_responses.column_dimensions['F'].width = 50
        ws_responses.column_dimensions['G'].width = 15
        ws_responses.column_dimensions['H'].width = 30
        ws_responses.column_dimensions['I'].width = 10
        ws_responses.column_dimensions['J'].width = 18
        
        # Hoja 3: Divergencias
        ws_divergence = wb.create_sheet("Divergencias")
        
        div_headers = ["Pregunta", "Severidad", "Usuarios Involucrados", "Respuestas", "Varianza", "Estado"]
        for col, header in enumerate(div_headers, 1):
            cell = ws_divergence.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border
        
        for row, alert in enumerate(report["divergence_alerts"], 2):
            users = ", ".join([r["user_name"] for r in alert["responses_data"]])
            responses = " | ".join([f"{r['user_name']}: {r['answer']}" for r in alert["responses_data"]])
            status = "Resuelto" if alert["is_resolved"] else "Pendiente"
            
            # Color por severidad
            severity_colors = {
                "critical": "FF0000",
                "high": "FF6B6B",
                "medium": "FFB347",
                "low": "77DD77"
            }
            severity_fill = PatternFill(
                start_color=severity_colors.get(alert["severity"], "FFFFFF"),
                end_color=severity_colors.get(alert["severity"], "FFFFFF"),
                fill_type="solid"
            )
            
            ws_divergence.cell(row=row, column=1, value=alert["question_text"]).border = border
            cell_sev = ws_divergence.cell(row=row, column=2, value=alert["severity"].upper())
            cell_sev.fill = severity_fill
            cell_sev.border = border
            ws_divergence.cell(row=row, column=3, value=users).border = border
            ws_divergence.cell(row=row, column=4, value=responses).border = border
            ws_divergence.cell(row=row, column=5, value=alert["variance"] or "N/A").border = border
            ws_divergence.cell(row=row, column=6, value=status).border = border
        
        ws_divergence.column_dimensions['A'].width = 50
        ws_divergence.column_dimensions['B'].width = 12
        ws_divergence.column_dimensions['C'].width = 30
        ws_divergence.column_dimensions['D'].width = 50
        ws_divergence.column_dimensions['E'].width = 12
        ws_divergence.column_dimensions['F'].width = 12
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
